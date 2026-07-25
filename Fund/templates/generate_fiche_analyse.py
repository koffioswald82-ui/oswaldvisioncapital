#!/usr/bin/env python3
"""
OVC — Générateur de la fiche de travail "Analyse & Score de conviction"
Usage : python3 generate_fiche_analyse.py
Sortie : template-fiche-analyse-conviction.xlsx (dans le même dossier)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

# ── Palette (cohérente avec le site : ink navy / gold / body) ──────────
INK      = '1A1A2E'
GOLD     = 'C9832A'
GOLD_BG  = 'FEF3E2'
CREAM    = 'F5F2EC'
LINE     = 'E8E4F0'
GREEN    = '1A7A4A'
GREEN_BG = 'E8F5EE'
RED      = 'C0392B'
RED_BG   = 'FDE8E8'
WHITE    = 'FFFFFF'
BODY     = '4A4A68'

HEADER_FONT   = Font(name='Calibri', size=11, bold=True, color=WHITE)
HEADER_FILL   = PatternFill('solid', fgColor=INK)
TITLE_FONT    = Font(name='Calibri', size=16, bold=True, color=INK)
SUB_FONT      = Font(name='Calibri', size=10, italic=True, color=BODY)
LABEL_FONT    = Font(name='Calibri', size=10, bold=True, color=INK)
GOLD_FONT     = Font(name='Calibri', size=10, bold=True, color=GOLD)
THIN          = Side(style='thin', color=LINE)
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

def style_header_row(ws, row, first_col, last_col, height=22):
    ws.row_dimensions[row].height = height
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical='center', horizontal='left', indent=1)

def title_block(ws, text, sub, span_cols=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span_cols)
    c2 = ws.cell(row=2, column=1, value=sub)
    c2.font = SUB_FONT
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16

# ════════════════════════════════════════════════════════════
# FEUILLE 1 — SCORE DE CONVICTION
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Score de conviction'
title_block(ws1, 'Score de conviction', 'Oswald Vision Capital — grille de notation pondérée, à remplir avant publication')

headers = ['Critère', 'Poids (%)', 'Note (1-10)', 'Contribution', 'Commentaire']
for i, h in enumerate(headers, start=1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 1, 5)

criteria = [
    ('Qualité du business (moat, marges, position concurrentielle)', 25),
    ('Valorisation (P/E, EV/EBITDA vs historique et secteur)', 20),
    ('Catalyseurs à 12 mois (résultats, lancement, régulation...)', 15),
    ('Risques identifiés (marché, régulation, exécution)', 15, True),  # True = criterion scored inversely (higher risk = lower score contribution) -- handled by comment only
    ('Momentum / technique (tendance, volumes)', 10),
    ('Qualité et fraîcheur des données disponibles', 10),
    ('Alignement avec la stratégie du fonds (zone, style)', 5),
]

row = 5
for crit in criteria:
    name, weight = crit[0], crit[1]
    ws1.cell(row=row, column=1, value=name).border = BORDER
    ws1.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='center')
    wcell = ws1.cell(row=row, column=2, value=weight/100)
    wcell.number_format = '0%'
    wcell.border = BORDER
    wcell.alignment = Alignment(horizontal='center')
    ncell = ws1.cell(row=row, column=3)
    ncell.border = BORDER
    ncell.alignment = Alignment(horizontal='center')
    ncell.fill = PatternFill('solid', fgColor=CREAM)
    contrib = ws1.cell(row=row, column=4, value=f'=B{row}*C{row}')
    contrib.number_format = '0.00'
    contrib.border = BORDER
    contrib.alignment = Alignment(horizontal='center')
    ws1.cell(row=row, column=5).border = BORDER
    row += 1

last_crit_row = row - 1
total_row = row + 1
ws1.cell(row=total_row, column=1, value='SCORE PONDÉRÉ TOTAL (/10)').font = LABEL_FONT
total_cell = ws1.cell(row=total_row, column=4, value=f'=SUM(D5:D{last_crit_row})')
total_cell.number_format = '0.00'
total_cell.font = Font(bold=True, size=13, color=INK)
total_cell.fill = PatternFill('solid', fgColor=GOLD_BG)
total_cell.border = BORDER

reco_row = total_row + 1
ws1.cell(row=reco_row, column=1, value='RECOMMANDATION SUGGÉRÉE').font = LABEL_FONT
reco_cell = ws1.cell(
    row=reco_row, column=4,
    value=f'=IF(D{total_row}>=7.5,"BUY",IF(D{total_row}>=5.5,"HOLD",IF(D{total_row}>=3.5,"NEUTRAL","SELL")))'
)
reco_cell.font = Font(bold=True, size=12)
reco_cell.alignment = Alignment(horizontal='center')
reco_cell.border = BORDER

# Mise en forme conditionnelle sur la recommandation (cohérent avec les badges du site : BUY vert, SELL rouge)
ws1.conditional_formatting.add(
    reco_cell.coordinate,
    CellIsRule(operator='equal', formula=['"BUY"'], fill=PatternFill('solid', fgColor=GREEN_BG), font=Font(color=GREEN, bold=True))
)
ws1.conditional_formatting.add(
    reco_cell.coordinate,
    CellIsRule(operator='equal', formula=['"SELL"'], fill=PatternFill('solid', fgColor=RED_BG), font=Font(color=RED, bold=True))
)
ws1.conditional_formatting.add(
    reco_cell.coordinate,
    CellIsRule(operator='equal', formula=['"HOLD"'], fill=PatternFill('solid', fgColor=GOLD_BG), font=Font(color=GOLD, bold=True))
)

# Échelle de couleur sur les notes individuelles
ws1.conditional_formatting.add(
    f'C5:C{last_crit_row}',
    ColorScaleRule(start_type='num', start_value=1, start_color='F8D7D3',
                    mid_type='num', mid_value=5.5, mid_color='FEF3E2',
                    end_type='num', end_value=10, end_color='D9F0E1')
)

note = ws1.cell(row=reco_row + 2, column=1,
    value="Note : la note (colonne C) est saisie manuellement pour chaque critère, de 1 (très faible) à 10 (excellent). "
          "Pour le critère \"Risques\", noter 10 = risques bien maîtrisés/limités, 1 = risques élevés non couverts. "
          "Le seuil BUY/HOLD/NEUTRAL/SELL ci-dessus est indicatif — la décision finale reste éditoriale.")
note.font = SUB_FONT
note.alignment = Alignment(wrap_text=True)
ws1.merge_cells(start_row=reco_row + 2, start_column=1, end_row=reco_row + 2, end_column=5)
ws1.row_dimensions[reco_row + 2].height = 45

dv_note = DataValidation(type='whole', operator='between', formula1=1, formula2=10, showErrorMessage=True,
                          errorTitle='Note invalide', error='Entrez un nombre entier entre 1 et 10.')
ws1.add_data_validation(dv_note)
dv_note.add(f'C5:C{last_crit_row}')

ws1.column_dimensions['A'].width = 46
ws1.column_dimensions['B'].width = 11
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 34
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A5'

# ════════════════════════════════════════════════════════════
# FEUILLE 2 — DONNÉES CLÉS (mappable directement sur le CMS admin)
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Données clés')
title_block(ws2, 'Données clés', 'Champs directement transférables dans le CMS (admin-ovc-secret.html)', span_cols=3)

fields_sheet2 = [
    ('Identification', None),
    ('Titre de l\'analyse', ''),
    ('Ticker', ''),
    ('Zone géographique', ''),  # dropdown: africa/americas/europe/asia
    ('Stratégie', ''),          # dropdown: growth/value/blend
    ('Secteur', ''),
    ('Auteur', 'Oswald Jaûres KOFFI'),
    ('Valorisation & cible', None),
    ('Cours actuel', ''),
    ('Devise', ''),
    ('Cours cible 12 mois', ''),
    ('Upside / downside (%)', '=IFERROR((C11-C10)/C10,"")'),
    ('P/E (x)', ''),
    ('EV/EBITDA (x)', ''),
    ('ROE (%)', ''),
    ('FCF Yield (%)', ''),
    ('Recommandation (BUY/HOLD/NEUTRAL/SELL)', ''),
    ('Méta article', None),
    ('Lede / accroche (1-2 phrases)', ''),
    ('Tags (séparés par des virgules)', ''),
    ('Temps de lecture estimé (min)', ''),
]

r = 4
section_rows = []
for item in fields_sheet2:
    label, default = item
    if default is None:
        section_rows.append(r)
        ws2.cell(row=r, column=1, value=label)
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cell = ws2.cell(row=r, column=1)
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=GOLD)
        ws2.row_dimensions[r].height = 20
        r += 1
        continue
    lc = ws2.cell(row=r, column=1, value=label)
    lc.font = LABEL_FONT
    lc.border = BORDER
    vc = ws2.cell(row=r, column=2, value=default if not str(default).startswith('=') else None)
    if str(default).startswith('='):
        ws2.cell(row=r, column=2, value=default)
    vc.border = BORDER
    vc.fill = PatternFill('solid', fgColor=CREAM)
    if 'Upside' in label:
        ws2.cell(row=r, column=2).number_format = '+0.0%;-0.0%'
    r += 1

# Dropdowns cohérents avec les valeurs utilisées dans le code du site
dv_zone = DataValidation(type='list', formula1='"africa,americas,europe,asia"', showErrorMessage=True)
ws2.add_data_validation(dv_zone)
dv_zone.add('B7')

dv_strat = DataValidation(type='list', formula1='"growth,value,blend"', showErrorMessage=True)
ws2.add_data_validation(dv_strat)
dv_strat.add('B8')

dv_reco = DataValidation(type='list', formula1='"BUY,HOLD,NEUTRAL,SELL"', showErrorMessage=True)
ws2.add_data_validation(dv_reco)
dv_reco.add('B17')

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 28
ws2.sheet_view.showGridLines = False

# ════════════════════════════════════════════════════════════
# FEUILLE 3 — THÈSE D'INVESTISSEMENT
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Thèse d\'investissement')
title_block(ws3, 'Thèse d\'investissement', 'Structure de rédaction — à recopier dans les sections de l\'article', span_cols=2)

thesis_fields = [
    'Thèse haussière (bull case) — 3 points clés',
    'Thèse baissière (bear case) — 3 points clés',
    'Catalyseurs identifiés (événements à surveiller)',
    'Risques principaux',
    'Comparables sectoriels (peers)',
    'Sources utilisées (rapports, données, liens)',
    'Conclusion / verdict éditorial',
]
r = 4
for label in thesis_fields:
    lc = ws3.cell(row=r, column=1, value=label)
    lc.font = LABEL_FONT
    lc.fill = PatternFill('solid', fgColor=INK)
    lc.font = Font(bold=True, color=WHITE, size=10)
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    ws3.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=2)
    box = ws3.cell(row=r, column=1)
    box.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    box.border = BORDER
    box.fill = PatternFill('solid', fgColor=CREAM)
    ws3.row_dimensions[r].height = 20
    r += 4

ws3.column_dimensions['A'].width = 60
ws3.column_dimensions['B'].width = 20
ws3.sheet_view.showGridLines = False

# ════════════════════════════════════════════════════════════
# FEUILLE 4 — CHECKLIST AVANT PUBLICATION
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Checklist publication')
title_block(ws4, 'Checklist avant publication', 'À valider avant de publier ou planifier l\'article dans le CMS', span_cols=3)

checklist = [
    'Score de conviction complété (feuille 1)',
    'Toutes les données clés remplies (feuille 2)',
    'Thèse haussière ET baissière rédigées (pas seulement le cas favorable)',
    'Sources vérifiées et citées',
    'Chiffres recoupés avec au moins 2 sources',
    'Disclaimer / avertissement risque présent',
    'Version anglaise (data-en) rédigée si publication bilingue',
    'Ticker et zone corrects dans le CMS',
    'Slug généré et vérifié (pas de doublon)',
    'Aperçu (cmsPreview) relu avant publication',
]
r = 4
headers4 = ['Vérification', 'Fait ?', 'Note']
for i, h in enumerate(headers4, start=1):
    ws4.cell(row=r, column=i, value=h)
style_header_row(ws4, r, 1, 3)
r += 1
dv_check = DataValidation(type='list', formula1='"☐ À faire,✓ Fait"', showErrorMessage=True)
ws4.add_data_validation(dv_check)
for item in checklist:
    ws4.cell(row=r, column=1, value=item).border = BORDER
    ws4.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='center')
    chk = ws4.cell(row=r, column=2, value='☐ À faire')
    chk.border = BORDER
    chk.alignment = Alignment(horizontal='center')
    dv_check.add(chk.coordinate)
    ws4.cell(row=r, column=3).border = BORDER
    r += 1

ws4.conditional_formatting.add(
    f'B5:B{r-1}',
    CellIsRule(operator='equal', formula=['"✓ Fait"'], fill=PatternFill('solid', fgColor=GREEN_BG), font=Font(color=GREEN, bold=True))
)

ws4.column_dimensions['A'].width = 55
ws4.column_dimensions['B'].width = 14
ws4.column_dimensions['C'].width = 30
ws4.sheet_view.showGridLines = False

wb.save('template-fiche-analyse-conviction.xlsx')
print('Fichier généré : template-fiche-analyse-conviction.xlsx')
